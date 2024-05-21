"""Main module."""
import os
import logging
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from rvtools2azmigrate.config import VINFO_SHEET_NAME, STORAGE_COLUMN_NAME
from rvtools2azmigrate.rvtools_vm import RvToolsVM

log = logging.getLogger(__name__)


def get_column_data_by_name(ws, column_name: str, default=None):
    """Get column data by name

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Worksheet
        column_name (str): Column name

    Returns:
        list: List of values or None if not found
    """
    for column_cell in ws.iter_cols(1, ws.max_column):  # iterate column cell
        if column_cell[0].value == column_name:  # check for your column
            return [data.value for data in column_cell[1:]]  # iterate your column
    log.warning(f"Column {column_name} not found")
    return None


def get_rvtools_vms(rvtools_data: dict, anonymized: bool, filter_off_vms: bool) -> list():
    """Get a list of RvToolsVM objects based on the RVTools data

    Args:
        rvtools_data (dict): RVTools data as a dict
        anonymized (bool): Anonymize the output data ?
        filter_off_vms (bool): Filter the powered-off VMs

    Returns:
        list(): List of RvToolsVM objects
    """
    rvtools_vms = []
    nb_vms = len(rvtools_data["vms_name"])
    vm_name_already_used = []
    for i in range(nb_vms):
        rv_vm = RvToolsVM(
            is_anonymized=anonymized,
            name=rvtools_data["vms_name"][i],
            power_state=rvtools_data["power_states"][i],
            cores=rvtools_data["cores"][i],
            memory=rvtools_data["memory"][i],
            os_config=rvtools_data["os_config"][i],
            os_vmtools=rvtools_data["os_vmtools"][i],
            storage_capacity=rvtools_data["storage_capacity"][i],
            is_mib=rvtools_data.get("is_mib", True),
            dns_name=rvtools_data["dns_name"][i],
            uuid=rvtools_data["uuid"][i],
            firmware=rvtools_data["firmware"][i],
        )
        # Filtering powered off VMs
        if rv_vm.power_state == "poweredOff" and filter_off_vms:
            _action = "Ignored"
        else:
            _action = "Processed"
            rvtools_vms.append(rv_vm)
        log.debug(f"{_action} VM {rv_vm.name} -> {rv_vm.power_state} ({i+1}/{nb_vms})")
    return rvtools_vms


def write_azmigrate_file(rvtools_vms: list, output: str):
    """Write Azure Migrate file

    Args:
        rvtools_vms (list): List of RvToolsVM
        output (str): Output file path
    """
    vm_name_already_used = []
    with open(output, "w", newline="") as f:
        # Build CSV headers based on data keys
        _first_vm = rvtools_vms[0].convert_to_az_migrate()
        writer = csv.DictWriter(f, fieldnames=_first_vm.as_csv_row().keys())
        writer.writeheader()
        # Append VM data
        for rv_vm in rvtools_vms:
            # Get the AZ migrate data from RVTools data
            vm_data = rv_vm.convert_to_az_migrate()
            # Check name uniqueness
            vm_name = vm_data.server_name
            if vm_name in vm_name_already_used:
                # Add an index to VM with similar naming
                vm_name = f"{vm_name}-{vm_name_already_used.count(vm_name)}"
            writer.writerow(vm_data.as_csv_row(forced_name=vm_name))
            # Save name occurrence (for uniqueness checks)
            vm_name_already_used.append(vm_data.server_name)


def parse_rvtools_data(rvtools: str) -> dict:
    """Parse the RVTools data to export VMs details

    Args:
        rvtools (str): Input file in RVTools format

    Returns:
        dict: RVTools data per column
    """
    # Open workbook and vInfo sheet
    wb = load_workbook(filename=rvtools, data_only=True)
    ws = wb[VINFO_SHEET_NAME]

    # Lookup for the storage column name based on MiB or MB unit
    is_mib = True
    storage_column_name = STORAGE_COLUMN_NAME + " MiB"
    storage_data = get_column_data_by_name(ws, column_name=storage_column_name)
    if not storage_data:
        is_mib = False
        storage_column_name = STORAGE_COLUMN_NAME + " MB"
        storage_data = get_column_data_by_name(ws, column_name=storage_column_name)
    log.info(f"Using the storage column name: {storage_column_name}")

    # Returns useful columns data
    return {
        "vms_name": get_column_data_by_name(ws, column_name="VM"),
        "power_states": get_column_data_by_name(ws, column_name="Powerstate"),
        "cores": get_column_data_by_name(ws, column_name="CPUs"),
        "memory": get_column_data_by_name(ws, column_name="Memory"),
        "os_config": get_column_data_by_name(ws, column_name="OS according to the configuration file"),
        "os_vmtools": get_column_data_by_name(ws, column_name="OS according to the VMware Tools"),
        "storage_capacity": get_column_data_by_name(ws, column_name=storage_column_name),
        "dns_name": get_column_data_by_name(ws, column_name="DNS Name"),
        "uuid": get_column_data_by_name(ws, column_name="VM UUID"),
        "firmware": get_column_data_by_name(ws, column_name="Firmware"),
        "is_mib": is_mib,
    }


def convert_rvtools_to_azmigrate(rvtools: str, output: str, anonymized: bool, filter_off_vms: bool):
    """Convert RVTools file to Azure Migrate format

    Args:
        rvtools (str): Input file in RVTools format
        output (str): Output file in Azure Migrate CSV format
        anonymized (bool): Anonymize VM names
        filter_off_vms (bool): Filter the powered-off VMs
    """
    # Read data from RVTools file
    rvtools_data = parse_rvtools_data(rvtools)
    log.info(f"We have found {len(rvtools_data['vms_name'])} VMs in the file {rvtools}")

    # Create a list of VMs
    rvtools_vms = get_rvtools_vms(rvtools_data, anonymized, filter_off_vms)
    log.info(f"VM data set from RVtools was built successfully")

    # Export VMs to an AZ migrate CSV file
    write_azmigrate_file(rvtools_vms, output)
    log.info(f"File {output} created successfully with {len(rvtools_vms)} VMs")
    return 0
