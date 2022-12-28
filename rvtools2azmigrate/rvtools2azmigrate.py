"""Main module."""
import os
import click
import logging
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

VINFO_SHEET_NAME = "vInfo"
STORAGE_COLUMN_NAME = "Provisioned"
DEFAULT_OS_NAME = "Windows Server 2019 Datacenter"


def get_column_data_by_name(ws, column_name: str):
    """Get column data by name

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Worksheet
        column_name (str): Column name

    Returns:
        list: List of values
    """
    for column_cell in ws.iter_cols(1, ws.max_column):  # iterate column cell
        if column_cell[0].value == column_name:  # check for your column
            return [data.value for data in column_cell[1:]]  # iterate your column
    log.error(f"Column {column_name} not found")


def get_unique_vm_name(vm_name_already_used: list, dns_name: str, uuid: str, index: int) -> str:
    """Get unique VM name

    Args:
        vm_name_already_used (list): List of VM names already used
        dns_name (str): VM DNS name
        uuid (str): VM UUID
        index (int): VM index in the list

    Returns:
        str: Unique VM name
    """
    vm_name = uuid  # Default VM name is the UUID
    if dns_name:  # If DNS name is not empty, use it as VM name
        vm_name = dns_name
    if vm_name not in vm_name_already_used:  # If VM name is not already used, return it
        return vm_name
    else:
        # If VM name is already used, add index to the name
        log.warning(f"VM name {vm_name} already used, adding index to the name: {vm_name}-{index}")
        return f"{vm_name}-{index}"


def build_azmigrate_data(rvtools_data: dict, anonymized: bool, mib: bool) -> list():
    """Build Azure Migrate file

    Args:
        rvtools_data (dict): RVTools data as a dict
        anonymized (bool): Anonymize the output data
        mib (bool): RVTools file is in MiB instead of MB for disk and memory

    return:
        list(): data for the Azure Migrate file
    """
    azmigrate_data = []
    nb_vms = len(rvtools_data["vms_name"])
    vm_name_already_used = []
    for i in range(nb_vms):
        log.debug(f"Processing VM {rvtools_data['vms_name'][i]} ({i+1}/{nb_vms})")
        vm_data = {}
        # Get a unique VM name based on the DNS name or the UUID and the current index
        # If anonymized, use only the UUID as VM name.
        vm_unique_name = get_unique_vm_name(
            vm_name_already_used,
            rvtools_data["dns_name"][i] if not anonymized else "",
            rvtools_data["uuid"][i],
            i,
        )
        vm_data["*Server name"] = vm_unique_name
        vm_name_already_used.append(vm_unique_name)
        vm_data["IP addresses"] = ""
        vm_data["*Cores"] = rvtools_data["cores"][i]
        vm_data["*Memory (In MB)"] = rvtools_data["memory"][i]
        if rvtools_data["os_vmtools"][i]:
            vm_data["*OS name"] = rvtools_data["os_vmtools"][i]
        elif rvtools_data["os_config"][i]:
            vm_data["*OS name"] = rvtools_data["os_config"][i]
        else:
            log.warning(f"OS not found for VM {rvtools_data['vms_name'][i]}")
            vm_data["*OS name"] = DEFAULT_OS_NAME
        vm_data["OS version"] = ""
        if "64-bit" in vm_data["*OS name"]:
            vm_data["OS architecture"] = "x64"
        elif "32-bit" in vm_data["*OS name"]:
            vm_data["OS architecture"] = "x86"
        else:
            vm_data["OS architecture"] = ""
        vm_data["CPU utilization percentage"] = ""
        vm_data["Memory utilization percentage"] = ""
        vm_data["Network adapters"] = ""
        vm_data["Network In throughput"] = ""
        vm_data["Network Out throughput"] = ""
        if rvtools_data["firmware"][i] == "efi":
            vm_data["Boot type"] = "UEFI"
        else:
            vm_data["Boot type"] = "BIOS"
        vm_data["Number of disks"] = ""
        if mib:
            vm_data["Disk 1 size (In GB)"] = int((rvtools_data["storage_capacity"][i] / 1.04858) / 1024)
        else:
            vm_data["Disk 1 size (In GB)"] = int(rvtools_data["storage_capacity"][i] / 1024)
        vm_data["Disk 1 read throughput (MB per second)"] = ""
        vm_data["Disk 1 write throughput (MB per second)"] = ""
        vm_data["Disk 1 read ops (operations per second)"] = ""
        vm_data["Disk 1 write ops (operations per second)"] = ""
        vm_data["Disk 2 size (In GB)"] = ""
        vm_data["Disk 2 read throughput (MB per second)"] = ""
        vm_data["Disk 2 write throughput (MB per second)"] = ""
        vm_data["Disk 2 read ops (operations per second)"] = ""
        vm_data["Disk 2 write ops (operations per second)"] = ""
        azmigrate_data.append(vm_data)
    return azmigrate_data


def write_azmigrate_file(azmigrate_data: list, output: str):
    """Write Azure Migrate file

    Args:
        azmigrate_data (list): List of VMs data
        output (str): Output file path
    """
    with open(output, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=azmigrate_data[0].keys())
        writer.writeheader()
        for vm_data in azmigrate_data:
            writer.writerow(vm_data)


def convert_rvtools_to_azmigrate(rvtools: str, output: str, anonymized: bool, mib: bool):
    """Convert RVTools file to Azure Migrate format

    Args:
        rvtools (str): Input file in RVTools format
        output (str): Output file in Azure Migrate CSV format
        anonymized (bool): Anonymize VM names
        mib (bool): Use MiB instead of GB for storage and memory capacity
    """
    wb = load_workbook(filename=rvtools, data_only=True)
    ws = wb[VINFO_SHEET_NAME]

    if mib:
        log.debug("Using MiB for storage capacity")
        storage_column_name = STORAGE_COLUMN_NAME + " MiB"
    else:
        storage_column_name = STORAGE_COLUMN_NAME + " MB"
    log.info(f"Using the storage column name: {storage_column_name}")

    rvtools_data = {
        "vms_name": get_column_data_by_name(ws, column_name="VM"),
        "power_states": get_column_data_by_name(ws, column_name="Powerstate"),
        "cores": get_column_data_by_name(ws, column_name="CPUs"),
        "memory": get_column_data_by_name(ws, column_name="Memory"),
        "os_config": get_column_data_by_name(ws, column_name="OS according to the configuration file"),
        "os_vmtools": get_column_data_by_name(ws, column_name="OS according to the VMware Tools"),
        "storage_capacity": get_column_data_by_name(ws, column_name=storage_column_name),
        "dns_name": get_column_data_by_name(ws, column_name="DNS Name"),
        "uuid": get_column_data_by_name(ws, column_name="VM UUID"),
        "firmware": get_column_data_by_name(ws, column_name="Firmware"),
    }
    log.info(f"We have found {len(rvtools_data['vms_name'])} VMs in the file {rvtools}")
    azmigrate_data = build_azmigrate_data(rvtools_data, anonymized, mib)
    log.info(f"Data for Azure Migrate file built successfully")
    write_azmigrate_file(azmigrate_data, output)
    log.info(f"File {output} created successfully")
    return 0
